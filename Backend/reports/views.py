from rest_framework.viewsets import ViewSet
from websocket.models import *
from openpyxl import Workbook
from rest_framework.decorators import action ,permission_classes
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated, AllowAny

class ReporteViewSet(ViewSet):
    @action(detail=False, methods=['get'] , url_path='file', permission_classes=[IsAuthenticated]) 
    def generar_reporte_excel(self ,request): 
        wb = Workbook()
        ws_hotel = wb.create_sheet(title="Hotel - Registros")
        ws_hotel.append(['Usuario', 'consumo Total', 'presupuesto','consumo_desperdicio_total','eficiencia_energetica','kilo_vatio_hora_costo','fecha Actualización'])
        hotel_data = Hotel.objects.all()
    
        ws_nivel = wb.create_sheet(title="Consumo - Pisos")
        ws_nivel.append(['nivel', 'consumo','fecha_actualizacion'])
        nivel_data = Nivel.objects.all()

        ws_habitacion = wb.create_sheet(title="Consumo - Habitaciones")
        ws_habitacion .append(['numero ', 'consumo', 'nivel ','presencia_humana','temperatura','humedad','consumo_desperdicio','fecha_actualizacion']) 
        habitacion_data = Habitacion.objects.all()

        ws_dispositivos = wb.create_sheet(title="Consumo - Dispositivos")
        ws_dispositivos.append(['habitacion', 'tipo','consumo_actual','consumo_acumulado','estado_remoto','fecha_actualizacion '])
        dispositivos_data = Dispositivo.objects.all()

        ws_consumo = wb.create_sheet(title="RegistroConsumo - General")
        ws_consumo.append(['dispositivo', 'habitacion', 'consumo', 'estado_remoto', 'presencia_humana', 'temperatura', 'humedad', 'fecha'])
        consumo_data = RegistroConsumo.objects.all()

        ws_alerta = wb.create_sheet(title="Alertas - General")
        ws_alerta.append(['habitacion', 'tipo','mensaje','fecha_creacion'])
        alerta_data = Alerta.objects.all()


        # Escribir datos de Hoteles
        for item in hotel_data:
            fecha_actualizacion = item.fecha_actualizacion.replace(tzinfo=None) if item.fecha_actualizacion else None
            ws_hotel.append([item.user.username, item.consumo_total, item.presupuesto,item.consumo_desperdicio_total,item.eficiencia_energetica ,item.kilo_vatio_hora_costo, fecha_actualizacion])

        # Escribir datos de Niveles
        for item in nivel_data:
            fecha_actualizacion = item.fecha_actualizacion.replace(tzinfo=None) if item.fecha_actualizacion else None
            ws_nivel.append([item.nivel, item.consumo, fecha_actualizacion])

        # Escribir datos de Habitaciones
        for item in habitacion_data:
            fecha_actualizacion = item.fecha_actualizacion.replace(tzinfo=None) if item.fecha_actualizacion else None
            ws_habitacion.append([item.numero, item.consumo, item.nivel.nivel, item.presencia_humana, item.temperatura, item.humedad,item.consumo_desperdicio, fecha_actualizacion])

        # Escribir datos de Dispositivos
        for item in dispositivos_data:
            fecha_actualizacion = item.fecha_actualizacion.replace(tzinfo=None) if item.fecha_actualizacion else None
            ws_dispositivos.append([item.habitacion.numero, item.get_tipo_display(), item.consumo_actual,item.consumo_acumulado, item.estado_remoto, fecha_actualizacion])

        # Escribir datos de RegistroConsumo
        for item in consumo_data:
            fecha = item.fecha.replace(tzinfo=None) if item.fecha else None
            ws_consumo.append([item.dispositivo.get_tipo_display(),item.habitacion.numero,item.consumo, item.estado_remoto ,item.presencia_humana , item.temperatura , item.humedad, fecha])

        # Escribir datos de Alertas
        for item in alerta_data:
            fecha_creacion = item.fecha_creacion.replace(tzinfo=None) if item.fecha_creacion else None
            ws_alerta.append([item.habitacion.numero, item.get_tipo_display(),item.mensaje, fecha_creacion])

        # Eliminar la hoja por defecto (si no se usa)
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']


        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte_hotel_kamila.xlsx'

        wb.save(response)

        return response